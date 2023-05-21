import pandas as pd #pandas untuk membuat dataframe(df)
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
import json

class ExcelReportPlugin():
    def __init__(self, 
                 input_file, 
                 output_file):
        self.input_file = input_file
        self.output_file = output_file
    def main(self):
        df = self.read_input_file()
        df_transform = self.transform(df, 'Gender', 'Product line', 'Total', 'sum')
        self.create_output_file(df_transform)
        print('Workbook Created')
        wb = load_workbook(self.output_file)
        wb.active = wb['Report']
        f = open(r'C:\Users\PutuAndika\OneDrive - Migo\Desktop\Data Engineer Project\Bootcamp Digital Skola\project_1\automate_report\digitalskola\automate_report\configs\webhook.json')
        data = json.load(f)
        webhook_url = data['webhook_url']

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row
        
        self.barchart(wb.active, min_column, max_column, min_row, max_row)
        self.add_total(max_column, max_row, min_row, wb.active , 'Sales Report', '2019')
        self.save_file(wb)
        self.send_to_discord(wb, webhook_url)
    def read_input_file(self):
        df = pd.read_excel(self.input_file)
        print(df.head())
        return df
    def transform(self, df, index, columns, values, aggfunc):
        df_transform = df.pivot_table(index=index, 
                                      columns=columns, 
                                      values=values, 
                                      aggfunc=aggfunc).round()
        print(df_transform)
        return df_transform
    def create_output_file(self, df_transform):
        df_transform.to_excel(self.output_file, 
                sheet_name='Report', 
                startrow=4)
    def barchart(self, workbook, min_column, max_column, min_row, max_row):
        barchart = BarChart()

        data = Reference(workbook, 
                        min_col=min_column+1,
                        max_col=max_column,
                        min_row=min_row,
                        max_row=max_row
                        )

        categories = Reference(workbook,
                                min_col=min_column,
                                max_col=min_column,
                                min_row=min_row+1,
                                max_row=max_row
                                )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)
        workbook.add_chart(barchart, 'B12')
        barchart.title = 'Sales berdasarkan Produk'
        barchart.style = 2
    def add_total(self, max_column, max_row, min_row, workbook,workbook_title, workbook_subtitle):
        import string
        alphabet = list(string.ascii_uppercase)
        alphabet_excel = alphabet[:max_column]
        #[A,B,C,D,E,F,G]
        for i in alphabet_excel:
            if i != 'A':
                workbook[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
                workbook[f'{i}{max_row+1}'].style = 'Currency'

        workbook[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'
        
        workbook['A1'] = workbook_title
        workbook['A2'] = workbook_subtitle
        workbook['A1'].font = Font('Arial', bold=True, size=20)
        workbook['A2'].font = Font('Arial', bold=True, size=10)
    def save_file(self,workbook):
        workbook.save(self.output_file)
        print('File Saved')
    def send_to_discord(self, workbook, url):
        import discord
        from discord import SyncWebhook
        webhook = SyncWebhook.from_url(url)

        with open(file=self.output_file, mode='rb') as file:
            excel_file = discord.File(file)

        webhook.send('This is an automated report',
                     username='Sales Bot',
                     file=excel_file)
        print('Sent to discord')